# https://github.com/nblintao/douban-video-detector
# Author: Tao Lin (visor_bulgur_0i@icloud.com)

from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

# 我懒得做所有的爬取和验证工作，直接用的是豆伴/豆坟（https://blog.doufen.org/）的导出的xlsx备份文件
# 请在得到自己的这个文件后，改成对应的名字
orgin_filename = '豆伴(12345678).xlsx'

# 用你喜爱的浏览器打开https://movie.douban.com/subject/25754848/，复制Request Headers
header_str= '''
Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9
Accept-Encoding: gzip, deflate, br
Accept-Language: en,en-US;q=0.9,zh;q=0.8,zh-CN;q=0.7,zh-TW;q=0.6
Connection: keep-alive
Cookie: 这就是打个样，请把整段改成自己的Request Headers
Host: movie.douban.com
Sec-Fetch-Dest: document
Sec-Fetch-Mode: navigate
Sec-Fetch-Site: none
Sec-Fetch-User: ?1
Upgrade-Insecure-Requests: 1
User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 11_1_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36
'''
# Thanks to https://www.jianshu.com/p/8f66fadaf844
def str2dict(s,s1=';',s2='='):
    li=s.split(s1)
    res={}
    for kv in li:
        li2=kv.split(s2)
        if len(li2)>1:
            li2[0]=li2[0].replace(':','')
            res[li2[0]]=li2[1]
    return res
headers=str2dict(header_str,'\n',': ')
headers['accept-encoding'] = 'gzip'

# test_data = {
#     'https://movie.douban.com/subject/26926321/': ['咪咕视频', '爱奇艺视频', '腾讯视频', '1905电影网', '西瓜视频'],
#     'https://movie.douban.com/subject/30413128/': ['欢喜首映', '哔哩哔哩'],
#     'https://movie.douban.com/subject/2210031/': ['优酷视频'],
#     'https://movie.douban.com/subject/1295526/': ['优酷视频'],
#     'https://movie.douban.com/subject/1309226/': ['优酷视频', '腾讯视频', '1905电影网']
# }
def get_sites(url):
    # return test_data.get(url, [])
    r = requests.get(url, headers=headers)
    if not r.ok:
        print(r.status_code)
        return []
    soup = BeautifulSoup(r.text)
    ul = soup.find('ul', class_="bs")
    if ul is None:
        return []
    lis = ul.findChildren("li" , recursive=False)
    return [li.find('a').text.strip() for li in lis]

new_col = 9 # 1-based
site_to_col = {}
def fill_site(row_num, site):
    global new_col
    if site not in site_to_col:
        site_to_col[site] = new_col
        ws.cell(row=1, column=new_col).value = site
        print("add site " + site + " at " + str(new_col))
        new_col += 1
    ws.cell(row=row_num, column=site_to_col[site]).value = '✅'
    
wb = load_workbook(orgin_filename)
ws = wb['想看']
for row in ws.iter_rows(min_row=2):
    try:
        row_num = row[0].row
        video = row[0].value
        url = row[3].value
        sites = get_sites(url)
        print(video + " " + url + " " + str(sites))
        for site in sites:
            fill_site(row_num, site)
    except Exception as e:
        print(e)
        pass
filename = '新' + orgin_filename
wb.save(filename)
